#!/usr/bin/env python
import os
#import xlrd
import openpyxl

def getYear(worksheet):
    c0 = []
    c_ret = []
    idx = []
    idx_ret = []
    for i in range(worksheet.nrows):
        c0.append(worksheet.row_values(i)[0])
        idx.append(i)
    if 'Year' or 'year' in c0:
        for i, val in enumerate(c0):
            if type(val) == str:
                if val.isdigit():
                    val = int(val)
            elif type(val) == unicode:
                if val.isnumeric():
                    val = float(val)
            if (val > 1900) and (val < 2100):
                c_ret.append(val)
                idx_ret.append(i)
        return (c_ret, idx_ret)
    else:
        return ([], [])

def getValue(worksheet, idx):
    row = []
    for i in range(worksheet.ncols):
        row.append(worksheet.col_values(i)[idx])
    return row[1:]

def nameCol(row, name):
    col_names = []
    for i,val in enumerate(row):
        retStr = '\\' + name + '-' + str(i+1)
        col_names.append(retStr)
    return col_names
  
def writeRow(ws, row, offset, row_number):
    for col_index, col_value in enumerate(row):
        ws.cell(row=row_number, column=col_index+offset).value = col_value
    
for year_idx in range(2004,2014):
    year_idx = str(year_idx)
    print('============================\n Processing year ' + year_idx)
    path= '/home/nochi/PCA/python/'
    path += year_idx

    wb          = openpyxl.Workbook()
    ws          = wb.get_sheet_by_name('Sheet')


    # Write first column 
    for idx, val in enumerate(range(1978,2015)):
        ws.cell(row=1+idx, column=0).value = val
        

    sorted_list = sorted(os.listdir(path))
    new_list = []
    for file in sorted_list:
        if file.endswith("xls"):
            new_list.append(file)

    # Each Excel file
    offset = 1
    error_count = 0
    year_count = 0
    non_year_count = 0
    for idx, file in enumerate(new_list):
            book = openpyxl.load_workbook(filename = path + '/' + file)
            sheet = book['Sheet']
            #sheet = book.sheet_by_index(0)
            
            a1 = sheet.cell_value(rowx=0, colx=0)
            splitted_a1 = a1.rsplit()
            name = splitted_a1[0]
            years_of_sheet, idx = getYear(sheet)
            if idx:
                first_row = getValue(sheet, 0)
                col_names = nameCol(first_row, name)
                writeRow(ws, col_names, offset, 0)

                for n, year in enumerate(range(1978,2015)):
                    for k, year_of_sheet in enumerate(years_of_sheet):
                        if year_of_sheet == year:
                            row = getValue(sheet, idx[k])
                            writeRow(ws, row, offset, n+1)
                year_count += 1
                offset += len(col_names)

            else:
                non_year_count += 1

    print('In total ' + str(year_count + non_year_count) + ' variables.')
    print('Year variables: ' + str(year_count))
    print('Non year variables: ' + str(non_year_count))
    save_name = '../raw_workbook_' + year_idx + '.xlsx' 
    wb.save(save_name)

